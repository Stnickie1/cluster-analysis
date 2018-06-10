import openpyxl as xl
import copy
from UtilityClasses import *
from matplotlib import markers
from matplotlib.lines import Line2D
import Constants
from PyQt5.QtWidgets import QTableWidgetItem


def readExcelData(filename):
    columns = []
    workbook = xl.load_workbook(filename=filename)
    worksheet = workbook.get_active_sheet()
    columnData = []
    columnName = None
    for i, col in enumerate(worksheet.iter_cols(min_col=1, max_col=worksheet.max_column)):
        for cell in col:
            if cell.row == 1:
                columnName = cell.internal_value
            else:
                columnData.append(cell.internal_value)
        column = Column(columnName, columnData, i)
        columns.append(column)
        columnData = []
    return columns


def drawBrokenLine(points, canvas):
    figure = canvas.figure
    axes = figure.add_subplot(1, 1, 1)
    xpolygonData = []
    ypolygonData = []
    for point in points:
        xpolygonData.append(point.getX())
        ypolygonData.append(point.getY())
    if len(points) == 1:
        axes.plot(xpolygonData, ypolygonData, linestyle="None", marker="o", color='b', markersize=4)
    else:
        axes.plot(xpolygonData, ypolygonData, "b")
    canvas.draw()


def drawPolygon(polygon, canvas):
    newPolygonContainter = copy.deepcopy(polygon)
    lastPoint = Point(polygon[0].getX(), polygon[0].getY(), None)
    newPolygonContainter.append(lastPoint)
    drawBrokenLine(newPolygonContainter, canvas)


def crossingNumberAlgorithm(point, polygon):
    intersectionCounter = 0
    if len(polygon) < 3:
        return False
    A1 = 0
    B1 = 1
    C1 = - point.getY()
    for i in range(0, len(polygon)):
        if i == len(polygon) - 1:
            X1 = polygon[i].getX()
            Y1 = polygon[i].getY()
            X2 = polygon[0].getX()
            Y2 = polygon[0].getY()
        else:
            X1 = polygon[i].getX()
            Y1 = polygon[i].getY()
            X2 = polygon[i + 1].getX()
            Y2 = polygon[i + 1].getY()
        A2 = Y2 - Y1
        B2 = - (X2 - X1)
        C2 = (X2 - X1) * Y1 - (Y2 - Y1) * X1
        Y = point.getY()
        if A2 != 0:
            X = (- C2 - B2 * Y) / A2
            if Y1 > Y2:
                var = Y2
                Y2 = Y1
                Y1 = var
            if (X > point.getX()) and (Y < Y2) and (Y > Y1):
                intersectionCounter += 1
    if intersectionCounter % 2 == 0:
        return False
    else:
        return True


def getFlippedMarkerDictionary():
    newDict = {}
    for key, value in markers.MarkerStyle.markers.items():
        newDict[value] = key
    return newDict


def getSupportiveLine(columns, i, j):
    xData = columns[i].getData()
    yData = columns[j].getData()
    line = Line2D(xData, yData,
                  linestyle="None",
                  marker=Constants.DEFAULT_POINT_SHAPE,
                  color=Constants.INVISIBLE_COLOR,
                  markersize=Constants.DEFAULT_MARKER_SIZE_SMALL)
    return line


def fillTableWithData(tablewidget, data):
    tablewidget.setColumnCount(data.columnCount())
    tablewidget.setRowCount(data.rowCount())
    horizontalheaders = data.getColumnNames()
    verticalheaders = []
    rows = data.getRows()
    for i, row in enumerate(rows):
        verticalheaders.append(str(row.getIndex()))
        for j, element in enumerate(row):
            tablewidget.setItem(i, j, QTableWidgetItem(str(element)))
    tablewidget.setHorizontalHeaderLabels(horizontalheaders)
    tablewidget.setVerticalHeaderLabels(verticalheaders)


def fillTableWithCluster(tablewidget, cluster):
    if cluster.getSize() != 0:
        tablewidget.setColumnCount(len(cluster.getRows()[
                                           0].getDataArray()))  # TODO очень криво выходит. Нет адекватного способа узнать у кластера количество колонок и их имена.
        # Кластер должен содержать обьект типа Data вместо множества строк?
        tablewidget.setRowCount(cluster.getSize())
        verticalheaders = []
        rows = cluster.getRows()
        for i, row in enumerate(rows):
            verticalheaders.append(str(row.getIndex()))
            for j, element in enumerate(row):
                tablewidget.setItem(i, j, QTableWidgetItem(str(element)))
        tablewidget.setVerticalHeaderLabels(verticalheaders)


def calculate_ssw(cluster):
    res = 0
    if len(cluster) == 1 or len(cluster) == 0:
        return 0
    for i in range(0, len(cluster)):
        if i + 1 > len(cluster):
            return res
        else:
            elem = cluster[i]
            for j in range(i + 1, len(cluster)):
                for h in range(0, len(elem)):
                    res += pow(cluster[j][h] - elem[h], 2)
    return res


def calculate_sst(cluster):
    res = 0
    if len(cluster) == 1 or len(cluster) == 0:
        return 0
    for i in range(0, len(cluster)):
        if i + 1 > len(cluster):
            return res
        else:
            elem = cluster[i]
            for j in range(i + 1, len(cluster)):
                for h in range(0, len(elem)):
                    res += pow(cluster[j][h] - elem[h], 2)
    return res


def calculate_ssb(sst, ssw):
    return sst - ssw
